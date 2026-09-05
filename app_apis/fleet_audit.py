"""
app_apis.fleet_audit -- the three-way reconciliation behind the Fleet Audit page.

The question
------------
Three systems each hold a list of devices, and none of them agrees with the
others:

    ERPNext   `Customer Vehicle`  22,437 rows -- what we BELIEVE we installed
    IM        gps.im2m.ws          ~6,500 devices -- what IM is actually tracking
    Pilot     ksa.pilot-gps.com    per-account -- what Pilot is actually tracking

"Is this vehicle really on the platform the ERP says it is on?" has, until now,
been answerable only one vehicle at a time. This module answers it for the
whole fleet at once and stores the answer in `app_apis_fleet_audit`.

The join key
------------
`Customer Vehicle.device_serial` -- the IMEI. It earns the job: 19,851 distinct
values, ZERO duplicates, and 19,849 of them are 15-digit IMEIs. Both platforms
key their own rows by IMEI too, so no plate, name or customer ever has to be
matched -- which matters, because plates are spelled several ways across these
three systems and a fuzzy match that looks right is worse than no match.

ERP rows with no device serial cannot be reconciled at all. They are NOT
dropped: they appear with the verdict "No device serial", because 222 vehicles
marked Installed with no device recorded is itself a finding.

How each side is enumerated
---------------------------
ERP     one SQL read.

IM      `app_apis.im_connector.fetch_fleet()` -- ONE call. An empty filter body
        returns the entire fleet.

Pilot   ONE call, as the admin account: `GET /api/v3/vehicles` signed in with
        the Pilot Admin credentials. A partner login sees the whole estate, so
        one sign-in replaces ~1,800 of them and the audit drops from about 35
        minutes to a few seconds.

        `pilot_source="accounts"` still exists and signs in to every customer
        account in turn. It is NOT the default and should not be: it is thirty
        times slower, and it is also less correct -- an account whose password
        the ERP does not hold simply fails, and every vehicle behind it then
        reads as "Missing on Pilot" when Pilot has it perfectly well. It is
        kept for one case only: no admin credentials yet.

Why v3 and not the legacy `cmd=list`
------------------------------------
Both work. v3 wins on evidence: `cmd=list` answers `{"code":0,"msg":"OK"}` with
the rows under a `list` key and nothing under `data`, which is exactly the
shape that reads as "this account has no vehicles" to anyone not expecting it.
v3's `GET /api/v3/vehicles` returns them under `data` with `imei`, `veh_id`,
`vehiclenumber`, `folder`, `vin` and a `status` block. Verified on this
deployment: one customer account, 814 vehicles.

What this module does NOT do
----------------------------
It never writes to ERPNext, IM or Pilot. Every call is a read, and the only
table it writes is its own snapshot. An audit that could "fix" what it found
would be an audit nobody could trust to run unattended.
"""

import json
import time
from datetime import datetime, timedelta

import frappe
from frappe import _
from frappe.utils import cint, now_datetime

SNAPSHOT_DT = "app_apis_fleet_audit"
SIM_DT = "app_apis_sim_import"
SUMMARY_KEY = "app_apis_fleet_audit_summary"
PROGRESS_EVENT = "fleet_audit_progress"

READ_ROLES = ["System Manager", "Technical", "Support Team"]
RUN_ROLES = ["System Manager", "Technical"]

# Where a Pilot account may be recorded. Listing a field a site does not define
# is harmless -- the SQL builder below checks the meta first.
VEHICLE_EMAIL_FIELDS = (
	"email_pilot", "pilot_tow_email", "pilot_sfda_email",
	"pilot_tracking_only_email", "email_pilot2",
)
CUSTOMER_EMAIL_FIELDS = (
	"email_pilot", "email_pilot_tow", "email_pilot_sfda",
	"email_pilot_motorcycle", "email_pilot_tracking_only",
)

# One field may hold several addresses glued together -- the same separators
# app_apis.connector already has to cope with.
EMAIL_SEPARATORS = ("----", "---", "--", ",", ";", "|", "/", "\n", "\t", " ")

VERDICTS = (
	"OK",
	"Missing on Pilot",
	"Missing on IM",
	"Missing on both",
	"Deleted but still live",
	"Unexpected platform",
	"Not in ERP",
	"No device serial",
	"Not checked",
	"Unmatchable ID",
)


# --------------------------------------------------------------------------
# progress
# --------------------------------------------------------------------------


def _progress(stage: str, done: int = 0, total: int = 0, detail: str = "", user: str | None = None) -> None:
	"""Tell the page what is happening. Best-effort: never break a run over it."""
	try:
		frappe.publish_realtime(
			PROGRESS_EVENT,
			{
				"stage": stage,
				"done": done,
				"total": total,
				"percent": int(done * 100 / total) if total else 0,
				"detail": detail,
			},
			user=user or frappe.session.user,
		)
	except Exception:
		pass


# --------------------------------------------------------------------------
# side 1: the ERP
# --------------------------------------------------------------------------


def _erp_fleet() -> tuple[dict, list, dict]:
	"""(by_imei, without_imei, counts). One SQL read of `Customer Vehicle`."""
	meta = frappe.get_meta("Customer Vehicle")

	def has(fn):
		return bool(meta.get_field(fn))

	# Every platform flag this site actually defines. `ch_trakzee` is labelled
	# "IM Tracking" on the form -- the fieldname is historical, the meaning is
	# IM, and 5,766 vehicles carry it against only 1,821 with
	# server_type='TRAKZEE'. So the flag is the better signal of intent and
	# both are consulted.
	pilot_flags = [f for f in ("ch_pilot_wsl", "ch_pilot_tow", "ch_pilot_sfda",
	                           "ch_pilot_tracking_only") if has(f)]
	im_flags = [f for f in ("ch_trakzee",) if has(f)]

	# customer_name is deliberately absent here -- see the note below.
	cols = ["name", "device_serial", "license_plate", "customer",
	        "device_statues", "server_type", "subscription_expiry_date", "devices_type"]
	cols = [c for c in cols if has(c) or c == "name"]

	rows = frappe.db.sql(
		"select %s %s %s from `tabCustomer Vehicle`"
		% (
			", ".join("`%s`" % c for c in cols),
			(", " + ", ".join("`%s`" % f for f in pilot_flags)) if pilot_flags else "",
			(", " + ", ".join("`%s`" % f for f in im_flags)) if im_flags else "",
		),
		as_dict=True,
	)

	# The customer's NAME does not come from Customer Vehicle.
	#
	# There is a `customer_name` COLUMN on that table, but it is not a docfield
	# -- a leftover, filled on 3,419 of 22,437 rows. Reading it would put a name
	# on one vehicle in seven and leave the rest blank, which looks like missing
	# data rather than the wrong source. So the name is taken from `Customer`
	# itself: one query, then a dict lookup per row.
	#
	# The fallback is the Customer's docname, and on this site that IS the
	# name -- customers are named "مجموعه شركات الظاهري", not "CUST-0001" -- so
	# a customer with an empty customer_name still shows something meaningful.
	customer_names = dict(
		frappe.db.sql("select name, customer_name from `tabCustomer`") or []
	)

	by_imei: dict[str, dict] = {}
	without: list[dict] = []
	counts: dict[str, int] = {}

	for r in rows:
		status = (r.get("device_statues") or "").strip() or "(none)"
		counts[status] = counts.get(status, 0) + 1

		customer = r.get("customer")
		r["customer_name"] = (customer_names.get(customer) or customer) if customer else None

		server = (r.get("server_type") or "").strip()
		r["_expects_pilot"] = bool(
			any(cint(r.get(f)) for f in pilot_flags) or server.upper().startswith("PILOT")
		)
		r["_expects_im"] = bool(
			any(cint(r.get(f)) for f in im_flags) or server.upper() == "TRAKZEE"
		)

		imei = str(r.get("device_serial") or "").strip()
		if imei:
			# device_serial has no duplicates in this data, but never assume:
			# keep the first and remember that a second existed.
			if imei in by_imei:
				by_imei[imei].setdefault("_duplicates", []).append(r["name"])
			else:
				by_imei[imei] = r
		else:
			without.append(r)

	return by_imei, without, counts


# --------------------------------------------------------------------------
# side 2: IM
# --------------------------------------------------------------------------


def _im_fleet() -> tuple[dict, dict]:
	"""(by_imei, report). One call to IM."""
	from app_apis import im_connector

	started = time.monotonic()
	result = im_connector.fetch_fleet()
	elapsed = int((time.monotonic() - started) * 1000)

	if cint(result.get("code")) != 0:
		return {}, {
			"ok": False,
			"devices": 0,
			"elapsed_ms": elapsed,
			"error": "[%s] %s" % (result.get("code"), result.get("msg")),
		}

	rows = result.get("rows") or []
	by_imei = {}
	for row in rows:
		by_imei.setdefault(row["imei"], row)

	return by_imei, {
		"ok": True,
		"devices": len(by_imei),
		"rows_returned": len(rows),
		"elapsed_ms": elapsed,
		"account": (result.get("_im") or {}).get("account"),
	}


# --------------------------------------------------------------------------
# side 3: Pilot
# --------------------------------------------------------------------------


def _split_emails(raw) -> list[str]:
	cleaned = str(raw or "")
	for sep in EMAIL_SEPARATORS:
		cleaned = cleaned.replace(sep, " ")
	return [p.strip().lower() for p in cleaned.split() if "@" in p]


def pilot_accounts() -> list[tuple[str, int]]:
	"""Every distinct Pilot login the ERP knows, busiest first.

	Busiest first is not cosmetic: a capped or interrupted run then covers the
	most vehicles it can, instead of an arbitrary slice.
	"""
	weight: dict[str, int] = {}

	vmeta = frappe.get_meta("Customer Vehicle")
	for field in VEHICLE_EMAIL_FIELDS:
		if not vmeta.get_field(field):
			continue
		for raw, n in frappe.db.sql(
			"select `%s`, count(*) from `tabCustomer Vehicle` where ifnull(`%s`,'') <> '' group by `%s`"
			% (field, field, field)
		):
			for email in _split_emails(raw):
				weight[email] = weight.get(email, 0) + cint(n)

	cmeta = frappe.get_meta("Customer")
	for field in CUSTOMER_EMAIL_FIELDS:
		if not cmeta.get_field(field):
			continue
		for (raw,) in frappe.db.sql(
			"select `%s` from `tabCustomer` where ifnull(`%s`,'') <> ''" % (field, field)
		):
			for email in _split_emails(raw):
				weight.setdefault(email, 0)

	return sorted(weight.items(), key=lambda kv: -kv[1])


def _absorb_estate(by_imei: dict, rows, account, account_no: int = 1) -> int:
	"""Merge `pilot_admin.fetch_estate` rows into the by-IMEI map.

	The Administrator API keys devices by `uniqid`, which is the device serial
	but NOT always an IMEI: of 24,085 rows, 14,211 are 15-digit IMEIs and the
	rest are 10-digit registration ids and 23-25 digit SIM identifiers. They
	are all kept and joined on the raw value -- the non-IMEIs simply fail to
	match a Customer Vehicle, which is the correct outcome rather than a
	silently dropped device.

	`account_no` (1 or 2) is recorded in `pilot_accounts_no` on EVERY row this
	call touches, including one already merged in by a previous account --
	unlike the rest of the row's data, which is first-writer-wins, this set is
	the whole point of having two accounts: knowing WHICH one(s) actually see a
	given device, not just that at least one does.
	"""
	new = 0
	for row in rows or []:
		imei = str(row.get("imei") or "").strip()
		if not imei:
			continue
		if imei in by_imei:
			by_imei[imei].setdefault("pilot_accounts_no", set()).add(account_no)
			continue
		by_imei[imei] = {
			"imei": imei,
			"vehicle_no": row.get("vehicle_no"),
			"folder": row.get("folder"),
			"veh_id": row.get("veh_id"),
			"type": row.get("type"),
			"model": row.get("model"),
			"account": row.get("account") or account,
			"active": row.get("active"),
			"node_id": row.get("node_id"),
			"msisdn": row.get("msisdn"),
			"last_seen_epoch": row.get("last_seen_epoch"),
			"pilot_accounts_no": {account_no},
		}
		new += 1
	return new


def _pilot_fleet(source: str = "admin", max_accounts: int = 0,
                 user: str | None = None) -> tuple[dict, dict]:
	"""(by_imei, report). `source` is "admin" (default), "accounts" or "both".

	"admin" reads every ENABLED Pilot Administrator account -- one or two,
	whichever `app_apis` has configured -- and merges them into one map. Two
	accounts are two separate estates, not a primary and a backup, so a device
	only has to be on one of them to count as "on Pilot".
	"""
	from app_apis import connector, pilot_admin

	legacy = connector._settings()

	by_imei: dict[str, dict] = {}
	report = {
		"ok": True,
		"source": source,
		"accounts_total": 0,
		"accounts_tried": 0,
		"accounts_with_vehicles": 0,
		"accounts_auth_failed": 0,
		"accounts_errored": 0,
		"devices": 0,
		"admin_used": False,
		"admin_devices": 0,
		"admin_accounts_visible": None,
		"admin_runs": [],
		"failures": [],
		"elapsed_ms": 0,
	}
	started = time.monotonic()

	def absorb(rows, account):
		new = 0
		for row in rows or []:
			imei = str(row.get("imei") or "").strip()
			if not imei or imei in by_imei:
				continue
			status = row.get("status") or {}
			by_imei[imei] = {
				"imei": imei,
				"vehicle_no": str(row.get("vehiclenumber") or "").strip(),
				"folder": str(row.get("folder") or "").strip(),
				"veh_id": row.get("veh_id") or row.get("agentid"),
				"type": str(row.get("type") or "").strip(),
				"model": str(row.get("model") or row.get("configuration") or "").strip(),
				"vin": str(row.get("vin") or "").strip(),
				"account": account,
				"last_seen_epoch": cint(status.get("unixtimestamp")) or None,
			}
			new += 1
		return new

	# --- the admin account(s): ONE call each for the whole estate --------
	if source in ("admin", "both"):
		usable = [n for n in pilot_admin.enabled_admin_accounts()
		          if pilot_admin._is_usable(pilot_admin._settings(n))]

		if not usable:
			# Refuse rather than quietly fall back to the 1,800-sign-in sweep.
			# A silent fallback would turn a five-second audit into a
			# thirty-five-minute one with nothing on screen saying why.
			report["ok"] = False
			report["error"] = _(
				"No Pilot Admin connection is configured, so the fleet cannot be "
				"read in one call. Set the login and password in App APIs → Pilot "
				"Admin Connection (or Pilot Admin Connection 2) and tick it enabled."
			)
			report["elapsed_ms"] = int((time.monotonic() - started) * 1000)
			return {}, report

		any_ok = False
		for acc_no in usable:
			admin = pilot_admin._settings(acc_no)
			label = admin.get("label")

			_progress("pilot", 0, 1, _("Pilot ({0}): reading the whole estate in one call…").format(label), user)

			# The ADMINISTRATOR API, not v3. `cmd=vehicles` on the admin host
			# returns every device on the estate in a single request -- the
			# only Pilot call that can. v3's /vehicles is scoped to the
			# signed-in account and would need ~1,800 sign-ins to cover the
			# same ground.
			res = pilot_admin.fetch_estate(admin)
			run = {"account_no": acc_no, "label": label, "username": admin["username"]}

			if cint(res.get("code")) == 0:
				any_ok = True
				run["devices"] = _absorb_estate(by_imei, res.get("rows"), admin["username"], acc_no)
				run["attempts"] = (res.get("_pilot_admin") or {}).get("attempts")
				run["returned"] = (res.get("_pilot_admin") or {}).get("returned")

				# How many accounts that one login covers -- a cheap, reliable
				# call that puts the vehicle count in context.
				acc = pilot_admin.backend_accounts(admin)
				if cint(acc.get("code")) == 0:
					rows = acc.get("data") or []
					run["accounts_visible"] = len(rows)
					run["veh_count"] = sum(
						cint(a.get("veh_count")) for a in rows if isinstance(a, dict)
					)

				_progress("pilot", 1, 1,
				          _("Pilot ({0}): {1} devices").format(label, run["devices"]), user)
			else:
				run["error"] = "[%s] %s" % (res.get("code"), res.get("msg"))
				report["failures"].append({"account": admin["username"], "why": str(res.get("msg"))[:200]})

			report["admin_runs"].append(run)

		report["admin_used"] = any_ok
		report["admin_devices"] = sum(r.get("devices", 0) for r in report["admin_runs"])
		visible = [r["accounts_visible"] for r in report["admin_runs"] if "accounts_visible" in r]
		report["admin_accounts_visible"] = sum(visible) if visible else None
		report["devices"] = len(by_imei)

		if not any_ok:
			report["ok"] = False
			report["error"] = _("Every configured Pilot Admin connection failed: {0}").format(
				"; ".join(
					"%s: %s" % (r["label"], r["error"]) for r in report["admin_runs"] if r.get("error")
				)
			)
			if source == "admin":
				report["elapsed_ms"] = int((time.monotonic() - started) * 1000)
				return by_imei, report

	if source == "admin":
		report["devices"] = len(by_imei)
		report["elapsed_ms"] = int((time.monotonic() - started) * 1000)
		return by_imei, report

	# --- the fallback: every customer account, one sign-in each -----------
	accounts = pilot_accounts()
	report["accounts_total"] = len(accounts)
	if max_accounts:
		accounts = accounts[:max_accounts]

	total = len(accounts)
	for i, (email, _n) in enumerate(accounts, start=1):
		password = connector._password_for(email, legacy)
		if not password:
			report["accounts_errored"] += 1
			report["failures"].append({"account": email, "why": "no password available"})
			continue

		cfg = pilot_admin.settings_for(email, password)
		res = pilot_admin.request("/vehicles", settings=cfg)
		report["accounts_tried"] += 1

		code = cint(res.get("code"))
		if code == 0:
			found = absorb(res.get("data"), email)
			if found:
				report["accounts_with_vehicles"] += 1
		elif code == -401:
			report["accounts_auth_failed"] += 1
		else:
			report["accounts_errored"] += 1
			if len(report["failures"]) < 50:
				report["failures"].append({"account": email, "why": str(res.get("msg"))[:200]})

		if i % 10 == 0 or i == total:
			report["devices"] = len(by_imei)
			_progress(
				"pilot", i, total,
				_("Pilot: {0} of {1} accounts, {2} devices found").format(i, total, len(by_imei)),
				user,
			)

	report["devices"] = len(by_imei)
	report["elapsed_ms"] = int((time.monotonic() - started) * 1000)
	return by_imei, report


# --------------------------------------------------------------------------
# the reconciliation
# --------------------------------------------------------------------------


def _is_imei(key) -> bool:
	"""A 15-digit number. Pilot's `uniqid` often is not one.

	Of 24,085 devices the Administrator API returns, 14,211 carry a 15-digit
	IMEI; 7,578 carry a 10-digit registration id and ~2,300 carry a SIM-length
	number. Those can never match `Customer Vehicle.device_serial`, so counting
	them as "Not in ERP" would report 12,279 unknown devices where the real
	figure is 2,365. The distinction is the difference between a number
	somebody acts on and a number they learn to ignore.
	"""
	k = str(key or "").strip()
	return k.isdigit() and len(k) == 15


def _verdict(erp, on_pilot, on_im, pilot_checked=True, im_checked=True, key=None) -> tuple[str, list]:
	"""One verdict and every issue behind it.

	The verdict is the single worst thing true of this device; `issues` keeps
	all of them, because a device can be both missing on Pilot and unexpectedly
	present on IM and an operator needs to see both.

	`pilot_checked` / `im_checked` are not decoration. A platform that was not
	READ is not a platform the device is ABSENT from, and conflating the two is
	the single most dangerous thing a reconciliation can do: it would report
	17,000 vehicles as "Missing on Pilot" the moment a password expired, and
	somebody would act on it. An unchecked platform therefore produces no
	finding at all -- neither Missing nor Unexpected -- and a row whose only
	open question is that platform is called "Not checked", never "OK".
	"""
	issues = []

	if erp is None:
		if key is not None and not _is_imei(key):
			return "Unmatchable ID", [
				"The platform identifies this device as %r, which is not a 15-digit "
				"IMEI, so it cannot be matched against Customer Vehicle.device_serial. "
				"Not evidence of a missing ERP record." % str(key)[:40]
			]
		return "Not in ERP", [
			"Tracked by %s but no Customer Vehicle carries this device serial."
			% (" and ".join(x for x in (("Pilot" if on_pilot else None), ("IM" if on_im else None)) if x) or "a platform")
		]

	status = (erp.get("device_statues") or "").strip()
	expects_pilot = erp.get("_expects_pilot")
	expects_im = erp.get("_expects_im")
	live = status in ("Deleted", "Canceled Installation")

	if live and (on_pilot or on_im):
		where = " and ".join(x for x in (("Pilot" if on_pilot else None), ("IM" if on_im else None)) if x)
		issues.append("ERP says %s, but the device is still on %s." % (status, where))
		return "Deleted but still live", issues

	# A device the ERP does not consider installed, which no platform we
	# actually READ is tracking. If a platform went unread it could still be
	# there, so that is not a clean bill of health.
	unread = [n for n, checked in (("Pilot", pilot_checked), ("IM", im_checked)) if not checked]

	if status in ("Deleted", "Canceled Installation", "Processing"):
		if unread:
			issues.append("%s was not read in this run, so it cannot be confirmed that "
			              "this %s device is gone from %s."
			              % (" and ".join(unread), status.lower(), " and ".join(unread)))
			return "Not checked", issues
		return "OK", issues

	# From here on the ERP believes the device is Installed.
	missing_pilot = expects_pilot and pilot_checked and not on_pilot
	missing_im = expects_im and im_checked and not on_im

	if missing_pilot:
		issues.append("ERP expects this on Pilot; Pilot does not have it.")
	if missing_im:
		issues.append("ERP expects this on IM; IM does not have it.")

	if on_pilot and not expects_pilot:
		issues.append("On Pilot, but the ERP does not mark it as a Pilot vehicle.")
	if on_im and not expects_im:
		issues.append("On IM, but the ERP does not mark it as an IM vehicle.")

	if missing_pilot and missing_im:
		return "Missing on both", issues
	if missing_pilot:
		return "Missing on Pilot", issues
	if missing_im:
		return "Missing on IM", issues

	if (expects_pilot and not pilot_checked) or (expects_im and not im_checked):
		issues.append("The ERP expects this on %s, which was not read in this run."
		              % " and ".join(unread))
		return "Not checked", issues

	if not expects_pilot and not expects_im and not on_pilot and not on_im:
		if unread:
			issues.append("No platform is expected, and %s was not read." % " and ".join(unread))
			return "Not checked", issues
		issues.append("Installed, but no platform is expected and none has it.")
		return "Unexpected platform", issues

	if issues:
		return "Unexpected platform", issues

	if unread:
		issues.append("%s was not read in this run." % " and ".join(unread))
		return "Not checked", issues

	return "OK", issues


def _epoch_to_dt(epoch):
	"""Platform epoch -> a naive local datetime the desk will render.

	Wrapped in a try: both platforms have been seen sending 0, "" and, once, a
	timestamp far enough in the future to overflow. None of those is worth
	failing a 23,000-row audit over -- an empty "last seen" is honest.
	"""
	if not epoch:
		return None
	try:
		return datetime.fromtimestamp(int(epoch))
	except (ValueError, OverflowError, OSError, TypeError):
		return None


def _build_rows(erp_by_imei, erp_without, im_by_imei, pilot_by_imei, audited_at,
                pilot_checked=True, im_checked=True) -> tuple[list, dict]:
	"""Every reconciled row, plus the verdict tally."""
	tally: dict[str, int] = {v: 0 for v in VERDICTS}
	out = []

	keys = set(erp_by_imei) | set(im_by_imei) | set(pilot_by_imei)

	for imei in keys:
		erp = erp_by_imei.get(imei)
		im = im_by_imei.get(imei)
		pilot = pilot_by_imei.get(imei)
		verdict, issues = _verdict(erp, bool(pilot), bool(im), pilot_checked, im_checked, key=imei)
		tally[verdict] = tally.get(verdict, 0) + 1

		if erp and erp.get("_duplicates"):
			issues.append("Also on Customer Vehicle: %s" % ", ".join(erp["_duplicates"][:5]))

		out.append({
			"name": imei,
			"imei": imei,
			"erp_vehicle": (erp or {}).get("name"),
			"plate": (erp or {}).get("license_plate"),
			"customer": (erp or {}).get("customer"),
			"customer_name": (erp or {}).get("customer_name"),
			"erp_status": (erp or {}).get("device_statues") or "Not in ERP",
			"server_type": (erp or {}).get("server_type"),
			"device_model": (erp or {}).get("devices_type") or (im or {}).get("model") or (pilot or {}).get("model"),
			"expects_pilot": 1 if (erp or {}).get("_expects_pilot") else 0,
			"expects_im": 1 if (erp or {}).get("_expects_im") else 0,
			"on_pilot": 1 if pilot else 0,
			# Which admin account(s) actually saw this device -- not just that at
			# least one did. Two accounts are two separate estates (see
			# `_pilot_fleet`), so a device can be on one, the other, or both.
			"on_pilot_1": 1 if (pilot or {}).get("pilot_accounts_no", set()) & {1} else 0,
			"on_pilot_2": 1 if (pilot or {}).get("pilot_accounts_no", set()) & {2} else 0,
			"on_im": 1 if im else 0,
			"pilot_checked": 1 if pilot_checked else 0,
			"im_checked": 1 if im_checked else 0,
			"pilot_vehicle": (pilot or {}).get("vehicle_no"),
			"pilot_folder": (pilot or {}).get("folder"),
			"pilot_account": (pilot or {}).get("account"),
			"pilot_active": 1 if (pilot or {}).get("active") else 0,
			"pilot_last_seen": _epoch_to_dt((pilot or {}).get("last_seen_epoch")),
			"pilot_msisdn": (pilot or {}).get("msisdn") or None,
			"im_vehicle": (im or {}).get("vehicle_no"),
			"im_company": (im or {}).get("company"),
			"im_status": (im or {}).get("status"),
			"im_last_seen": _epoch_to_dt((im or {}).get("last_seen_epoch")),
			"subscription_expiry": (erp or {}).get("subscription_expiry_date"),
			"verdict": verdict,
			"issues": "\n".join(issues),
			# Filled in afterwards by `_apply_sim_data()`, from the last Lebara
			# SIM import -- not part of this reconciliation's own three-way
			# read, but merged onto the same row because it is one more fact
			# about the same device.
			"sim_status": None,
			"sim_last_seen": None,
			"sim_msisdn": None,
			"audited_at": audited_at,
		})

	# ERP rows with no device serial: reconcilable against nothing, and that is
	# the point of listing them.
	for erp in erp_without:
		status = (erp.get("device_statues") or "").strip()
		tally["No device serial"] = tally.get("No device serial", 0) + 1
		out.append({
			"name": "ERP-%s" % erp["name"],
			"imei": None,
			"erp_vehicle": erp["name"],
			"plate": erp.get("license_plate"),
			"customer": erp.get("customer"),
			"customer_name": erp.get("customer_name"),
			"erp_status": status or "(none)",
			"server_type": erp.get("server_type"),
			"device_model": erp.get("devices_type"),
			"expects_pilot": 1 if erp.get("_expects_pilot") else 0,
			"expects_im": 1 if erp.get("_expects_im") else 0,
			"on_pilot": 0,
			"on_pilot_1": 0,
			"on_pilot_2": 0,
			"on_im": 0,
			"pilot_checked": 1 if pilot_checked else 0,
			"im_checked": 1 if im_checked else 0,
			"pilot_vehicle": None, "pilot_folder": None, "pilot_account": None,
			"pilot_last_seen": None, "pilot_active": 0, "pilot_msisdn": None,
			"im_vehicle": None, "im_company": None, "im_status": None, "im_last_seen": None,
			"subscription_expiry": erp.get("subscription_expiry_date"),
			"verdict": "No device serial",
			"issues": "No device serial on the ERP record, so this vehicle cannot be "
			          "matched against Pilot or IM at all."
			          + ("" if status != "Installed" else " It is marked Installed."),
			"sim_status": None,
			"sim_last_seen": None,
			"sim_msisdn": None,
			"audited_at": audited_at,
		})

	return out, tally


COLUMNS = (
	"name", "imei", "erp_vehicle", "plate", "customer", "customer_name", "erp_status",
	"server_type", "device_model", "expects_pilot", "expects_im",
	"on_pilot", "on_pilot_1", "on_pilot_2", "on_im", "pilot_checked", "im_checked",
	"pilot_vehicle", "pilot_folder", "pilot_account", "pilot_last_seen", "pilot_active", "pilot_msisdn",
	"im_vehicle", "im_company", "im_status", "im_last_seen",
	"sim_status", "sim_last_seen", "sim_msisdn",
	"subscription_expiry", "verdict", "issues", "audited_at",
)


def _store(rows, audited_at) -> None:
	"""Replace the snapshot. Truncate + bulk insert: 23k rows in seconds.

	The ORM would take minutes for this and buy nothing -- these rows have no
	hooks, no links to maintain and no history worth keeping. The previous
	snapshot is not merged into: an audit is a photograph, and half of one
	photograph laid over another is not a record of anything.
	"""
	frappe.db.truncate(SNAPSHOT_DT)

	fields = ["name", "creation", "modified", "modified_by", "owner", "docstatus", "idx"] + [
		c for c in COLUMNS if c != "name"
	]
	user = frappe.session.user
	values = [
		[r["name"], audited_at, audited_at, user, user, 0, 0]
		+ [r.get(c) for c in COLUMNS if c != "name"]
		for r in rows
	]
	frappe.db.bulk_insert(SNAPSHOT_DT, fields=fields, values=values, chunk_size=5000)
	frappe.db.commit()


# --------------------------------------------------------------------------
# side 4: SIM data -- a Lebara export, uploaded by hand
# --------------------------------------------------------------------------
#
# Not a fourth platform this module reads live: nobody has given it an API,
# so the SIM status and last-connection date come from whatever .xlsx file
# somebody last uploaded. That file is parsed into its own table (`SIM_DT`)
# and merged onto the snapshot as two more columns -- exactly how Pilot's and
# IM's own "last seen" already sit on the same row, because a SIM's status is
# one more fact about the same device, not a separate audit.
#
# The merge survives a "Fetch All" refresh: `_store()` truncates and rebuilds
# the whole snapshot from a fresh ERP/Pilot/IM read, which would silently wipe
# the SIM columns along with everything else -- so `run_audit()` calls
# `_apply_sim_data()` again right after storing, re-applying the LAST upload
# onto the NEW snapshot. The SIM data itself lives in its own table precisely
# so it survives a snapshot rebuild it is not part of.


def _find_sim_header(ws) -> tuple[int | None, dict]:
	"""The row and column map for one worksheet, or (None, {}) if this is not it.

	Lebara's export can carry more than one sheet, and this module has no
	control over which is "the" data sheet -- so it looks for a header rather
	than assuming a name or position. A row counts once it has both MSISDN and
	IMEI headers, matched case-insensitively.
	"""
	for row in ws.iter_rows(min_row=1, max_row=10):
		values = {}
		for idx, c in enumerate(row):
			if c.value is not None:
				values[str(c.value).strip().lower()] = idx
		if "msisdn" in values and "imei" in values:
			return row[0].row, values
	return None, {}


def _coerce_datetime(v):
	"""A cell's date, in whatever form Lebara's export used.

	openpyxl hands back a real `datetime` for a date-formatted cell already;
	a plain string is only worth trying to parse, never worth failing an
	otherwise-good row over.
	"""
	if v in (None, ""):
		return None
	if isinstance(v, datetime):
		return v
	try:
		return frappe.utils.get_datetime(str(v).strip())
	except Exception:
		return None


def _parse_sim_workbook(path: str) -> list[dict]:
	"""Every usable row from a Lebara SIM export, keyed by column HEADER.

	Only five of the export's ~65 columns are read. A row with no MSISDN is
	dropped -- it is the identity of the SIM and the join key when no IMEI is
	present, so a row without one cannot be used for anything here.
	"""
	import openpyxl

	wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

	header_row, cols = None, {}
	target = None
	for ws in wb.worksheets:
		header_row, cols = _find_sim_header(ws)
		if header_row:
			target = ws
			break

	if not target:
		frappe.throw(
			_(
				"Could not find a header row with MSISDN and IMEI columns in this "
				"file. This importer expects a Lebara SIM export."
			),
			title=_("SIM Import"),
		)

	idx_msisdn = cols["msisdn"]
	idx_imei = cols["imei"]
	idx_iccid = cols.get("iccid")
	# The (M2M) variant is Lebara's per-line status; the plain column is a
	# secondary status field some exports omit. Either is usable; prefer the
	# M2M one when both are present.
	idx_status = cols.get("subscriber status (m2m)", cols.get("subscriber status"))
	idx_last = cols.get("last connection date")

	def cell(row, idx):
		if idx is None or idx >= len(row):
			return None
		v = row[idx].value
		return None if v in (None, "") else v

	out = []
	for row in target.iter_rows(min_row=header_row + 1):
		msisdn = str(cell(row, idx_msisdn) or "").strip()
		if not msisdn:
			continue
		out.append({
			"msisdn": msisdn,
			"imei": str(cell(row, idx_imei) or "").strip() or None,
			"iccid": str(cell(row, idx_iccid) or "").strip() or None,
			"sim_status": str(cell(row, idx_status) or "").strip() or None,
			"last_connection": _coerce_datetime(cell(row, idx_last)),
		})

	return out


def _apply_sim_data() -> dict:
	"""Merge the last uploaded SIM file onto the CURRENT snapshot.

	Matched by IMEI first -- the more trustworthy key, and the one most SIM
	rows carry. Rows still unmatched are tried again by MSISDN, against
	whatever MSISDN Pilot itself reported for that device (`pilot_msisdn`) --
	this is the fallback for the SIM rows that carry no IMEI at all.

	Every SIM column is reset before either pass, so a device the newest
	import no longer mentions does not keep showing a stale status from a
	previous upload.
	"""
	frappe.db.sql(
		"update `tab%s` set sim_status = NULL, sim_last_seen = NULL, sim_msisdn = NULL"
		% SNAPSHOT_DT
	)

	sim_rows = frappe.db.count(SIM_DT)
	if not sim_rows:
		frappe.db.commit()
		return {"sim_rows": 0, "matched_by_imei": 0, "matched_by_msisdn": 0}

	frappe.db.sql(
		"update `tab%s` fa join `tab%s` si on fa.imei = si.imei "
		"set fa.sim_status = si.sim_status, fa.sim_last_seen = si.last_connection, "
		"    fa.sim_msisdn = si.msisdn "
		"where ifnull(fa.imei, '') <> '' and ifnull(si.imei, '') <> ''"
		% (SNAPSHOT_DT, SIM_DT)
	)
	matched_imei = frappe.db._cursor.rowcount if frappe.db._cursor else 0

	frappe.db.sql(
		"update `tab%s` fa join `tab%s` si on fa.pilot_msisdn = si.msisdn "
		"set fa.sim_status = si.sim_status, fa.sim_last_seen = si.last_connection, "
		"    fa.sim_msisdn = si.msisdn "
		"where ifnull(fa.sim_status, '') = '' and ifnull(fa.pilot_msisdn, '') <> '' "
		"  and ifnull(si.msisdn, '') <> ''"
		% (SNAPSHOT_DT, SIM_DT)
	)
	matched_msisdn = frappe.db._cursor.rowcount if frappe.db._cursor else 0

	frappe.db.commit()
	return {"sim_rows": sim_rows, "matched_by_imei": matched_imei, "matched_by_msisdn": matched_msisdn}


# --------------------------------------------------------------------------
# the run
# --------------------------------------------------------------------------


def run_audit(include_im: int = 1, include_pilot: int = 1, pilot_source: str = "admin",
              max_accounts: int = 0, user: str | None = None) -> dict:
	"""The whole audit. Safe to call from a worker; writes only its own table."""
	user = user or frappe.session.user
	started = time.monotonic()
	audited_at = now_datetime()

	_progress("erp", 0, 0, _("Reading Customer Vehicle…"), user)
	erp_by_imei, erp_without, erp_counts = _erp_fleet()
	_progress("erp", 1, 1, _("{0} vehicles in the ERP").format(len(erp_by_imei) + len(erp_without)), user)

	im_by_imei, im_report = ({}, {"ok": None, "devices": 0, "skipped": True})
	if cint(include_im):
		_progress("im", 0, 1, _("Asking IM for the whole fleet…"), user)
		im_by_imei, im_report = _im_fleet()
		_progress("im", 1, 1, _("IM: {0} devices").format(im_report.get("devices", 0)), user)

	pilot_by_imei, pilot_report = ({}, {"ok": None, "devices": 0, "skipped": True})
	if cint(include_pilot):
		pilot_by_imei, pilot_report = _pilot_fleet(
			source=pilot_source, max_accounts=cint(max_accounts), user=user
		)

	# Was each platform actually READ? Only a platform that answered may produce
	# a "Missing on ..." finding -- see `_verdict`. A capped account sweep counts
	# as NOT read: it covers an arbitrary slice, and every vehicle behind an
	# un-swept account would otherwise be accused of being missing.
	im_checked = bool(cint(include_im)) and im_report.get("ok") is True
	pilot_checked = (
		bool(cint(include_pilot))
		and pilot_report.get("ok") is True
		and (
			pilot_report.get("source") == "admin"
			or not cint(max_accounts)
		)
	)
	im_report["checked"] = im_checked
	pilot_report["checked"] = pilot_checked

	_progress("reconcile", 0, 1, _("Reconciling…"), user)
	rows, tally = _build_rows(erp_by_imei, erp_without, im_by_imei, pilot_by_imei, audited_at,
	                          pilot_checked=pilot_checked, im_checked=im_checked)

	_progress("store", 0, 1, _("Storing {0} rows…").format(len(rows)), user)
	_store(rows, audited_at)

	# Re-apply the last SIM upload, if any -- `_store` just wiped the sim_*
	# columns along with the rest of the snapshot it rebuilt.
	sim_report = _apply_sim_data()

	summary = {
		"audited_at": str(audited_at),
		"duration_seconds": int(time.monotonic() - started),
		"run_by": user,
		"rows": len(rows),
		"erp": {
			"vehicles": len(erp_by_imei) + len(erp_without),
			"with_device_serial": len(erp_by_imei),
			"without_device_serial": len(erp_without),
			"by_status": erp_counts,
		},
		"im": im_report,
		"pilot": {k: v for k, v in pilot_report.items() if k != "failures"},
		"pilot_failures": (pilot_report.get("failures") or [])[:50],
		"sim": sim_report,
		"verdicts": tally,
		"matched": {
			"erp_and_pilot": sum(1 for r in rows if r["on_pilot"] and r["erp_vehicle"]),
			"erp_and_im": sum(1 for r in rows if r["on_im"] and r["erp_vehicle"]),
			"on_both_platforms": sum(1 for r in rows if r["on_pilot"] and r["on_im"]),
			"on_no_platform": sum(1 for r in rows if not r["on_pilot"] and not r["on_im"]),
		},
	}
	frappe.db.set_default(SUMMARY_KEY, json.dumps(summary, default=str))
	frappe.db.commit()

	_progress("done", 1, 1, _("Done: {0} rows").format(len(rows)), user)
	return summary


# --------------------------------------------------------------------------
# the whitelisted surface
# --------------------------------------------------------------------------


@frappe.whitelist()
def start_refresh(include_im: int = 1, include_pilot: int = 1, pilot_source: str = "admin",
                  max_accounts: int = 0) -> dict:
	"""Queue a run. Returns immediately; the page follows `fleet_audit_progress`.

	Still a background job even though the default path now takes seconds: the
	`accounts` fallback can take half an hour, and a request that sometimes
	returns in 5s and sometimes in 35 minutes is a request that will eventually
	time out in front of somebody. `timeout` is set explicitly because the
	`long` queue defaults to 25 minutes on this bench.
	"""
	frappe.only_for(RUN_ROLES)

	if pilot_source not in ("admin", "accounts", "both"):
		frappe.throw(_("Unknown Pilot source: {0}").format(pilot_source), title=_("Fleet Audit"))

	job = frappe.enqueue(
		"app_apis.fleet_audit.run_audit",
		queue="long",
		timeout=10800,
		include_im=cint(include_im),
		include_pilot=cint(include_pilot),
		pilot_source=pilot_source,
		max_accounts=cint(max_accounts),
		user=frappe.session.user,
		job_name="fleet_audit",
	)
	return {"queued": True, "job_id": getattr(job, "id", None)}


@frappe.whitelist()
def import_sim_file(file_url: str) -> dict:
	"""Ingest a Lebara SIM export (.xlsx) and merge it onto the snapshot.

	Runs in the request, not a background job: parsing is a few seconds even
	at 30,000 rows, and unlike the platform reads there is no network call
	here to time out.

	`file_url` is whatever Frappe's own upload endpoint (`frappe.ui.FileUploader`
	on the client) already saved -- this never accepts raw file bytes itself,
	so the upload goes through Frappe's existing, audited file-upload path and
	this only ever reads a file already on this server's disk.
	"""
	frappe.only_for(RUN_ROLES)

	file_url = str(file_url or "").strip()
	if not file_url:
		frappe.throw(_("file_url is required."), title=_("SIM Import"))

	file_doc = frappe.get_doc("File", {"file_url": file_url})
	rows = _parse_sim_workbook(file_doc.get_full_path())

	if not rows:
		frappe.throw(
			_("No usable rows found -- every row needs at least an MSISDN."),
			title=_("SIM Import"),
		)

	# Truncate + bulk insert: this is a replacement copy of Lebara's own
	# records, not something to merge row by row, and doing it this way is
	# seconds rather than minutes even at tens of thousands of rows.
	frappe.db.truncate(SIM_DT)
	now = now_datetime()
	user = frappe.session.user
	seen = set()
	values = []
	for r in rows:
		if r["msisdn"] in seen:
			continue
		seen.add(r["msisdn"])
		values.append([
			r["msisdn"], now, now, user, user, 0, 0,
			r["msisdn"], r["imei"], r["iccid"], r["sim_status"], r["last_connection"],
		])

	frappe.db.bulk_insert(
		SIM_DT,
		fields=["name", "creation", "modified", "modified_by", "owner", "docstatus", "idx",
		        "msisdn", "imei", "iccid", "sim_status", "last_connection"],
		values=values,
		chunk_size=5000,
	)
	frappe.db.commit()

	result = _apply_sim_data()
	result["file_rows"] = len(rows)
	result["duplicate_msisdns_dropped"] = len(rows) - len(values)
	return result


@frappe.whitelist()
def get_summary() -> dict:
	"""The last run's summary, or an empty shell before the first run."""
	frappe.only_for(READ_ROLES)

	raw = frappe.db.get_default(SUMMARY_KEY)
	summary = {}
	if raw:
		try:
			summary = json.loads(raw)
		except ValueError:
			summary = {}

	summary["snapshot_rows"] = frappe.db.count(SNAPSHOT_DT)
	summary["has_run"] = bool(raw)

	# The headline cards. Each is a QUESTION SOMEBODY ASKS, not a bucket:
	#
	#   "Need attention" used to be one tile summing every verdict that was not
	#   OK -- 10,722 rows that mixed 2,586 ERP records with no device serial in
	#   with 245 devices still being tracked after deletion. A number nobody can
	#   act on, because it is seven unrelated problems added together.
	#
	# `both_not_in_erp` is the one worth its own tile: a device that BOTH
	# platforms are tracking and no Customer Vehicle mentions. It is being
	# tracked, most likely billed, and invisible to the ERP.
	# Pilot's own device id (`uniqid`) is not always an IMEI -- of the roughly
	# 24,000 rows one estate sweep returns, only about 14,000 are real 15-digit
	# IMEIs; the rest are 10-digit registration numbers and SIM-length ids that
	# can never be a vehicle. Counting the raw sweep total as "devices on
	# Pilot" overstates the real fleet by that whole junk-id population --
	# every Pilot count below is therefore filtered to real IMEIs only.
	real_imei = "imei regexp '^[0-9]{15}$'"
	summary["cards"] = {
		"deleted_live": frappe.db.count(SNAPSHOT_DT, {"verdict": "Deleted but still live"}),
		"both_not_in_erp": cint(frappe.db.sql(
			"select count(*) from `tab%s` where on_pilot = 1 and on_im = 1"
			" and ifnull(erp_vehicle, '') = ''" % SNAPSHOT_DT
		)[0][0]),
		"missing_on_pilot": frappe.db.count(SNAPSHOT_DT, {"verdict": "Missing on Pilot"}),
		"no_device_serial": frappe.db.count(SNAPSHOT_DT, {"verdict": "No device serial"}),
		"pilot_wsl_devices": cint(frappe.db.sql(
			"select count(*) from `tab%s` where on_pilot_1 = 1 and %s" % (SNAPSHOT_DT, real_imei)
		)[0][0]),
		"pilot_2_devices": cint(frappe.db.sql(
			"select count(*) from `tab%s` where on_pilot_2 = 1 and %s" % (SNAPSHOT_DT, real_imei)
		)[0][0]),
		"pilot_devices": cint(frappe.db.sql(
			"select count(*) from `tab%s` where on_pilot = 1 and %s" % (SNAPSHOT_DT, real_imei)
		)[0][0]),
	}

	sim_row = frappe.db.sql(
		"select count(*), max(modified) from `tab%s`" % SIM_DT
	)
	summary["sim"] = {
		"rows": cint(sim_row[0][0]) if sim_row else 0,
		"last_import": str(sim_row[0][1]) if sim_row and sim_row[0][1] else None,
		"matched": frappe.db.sql(
			"select count(*) from `tab%s` where ifnull(sim_status, '') <> ''" % SNAPSHOT_DT
		)[0][0],
	}
	return summary


@frappe.whitelist()
def get_rows(verdict: str = "", erp_status: str = "", on_pilot: str = "", on_im: str = "",
             on_pilot_1: str = "", on_pilot_2: str = "",
             imei: str = "", vehicle: str = "", customer: str = "", sim_status: str = "",
             pilot_last_seen_from: str = "", pilot_last_seen_to: str = "",
             im_last_seen_from: str = "", im_last_seen_to: str = "",
             sim_last_seen_from: str = "", sim_last_seen_to: str = "",
             search: str = "", start: int = 0, limit: int = 100) -> dict:
	"""One page of the snapshot, filtered per column. Read-only.

	Written as explicit SQL rather than `frappe.get_all` because two of the
	column filters span two fields each -- Vehicle matches the plate OR the
	vehicle record, Customer matches the display name OR the customer id -- and
	Frappe's `or_filters` cannot express two independent OR groups ANDed
	together. Building the clause here keeps the filters exact and, more
	importantly, keeps the COUNT exact: the header count has to be the number of
	matching rows, not the number the page happened to fetch.

	Every value is bound as a parameter; no filter text reaches the SQL text.
	"""
	frappe.only_for(READ_ROLES)

	where, vals = [], []

	def like(term):
		return "%" + str(term).strip().replace("%", r"\%") + "%"

	if verdict:
		where.append("verdict = %s")
		vals.append(verdict)
	if erp_status:
		where.append("erp_status = %s")
		vals.append(erp_status)
	if str(on_pilot) in ("0", "1"):
		where.append("on_pilot = %s")
		vals.append(cint(on_pilot))
	if str(on_pilot_1) in ("0", "1"):
		where.append("on_pilot_1 = %s")
		vals.append(cint(on_pilot_1))
	if str(on_pilot_2) in ("0", "1"):
		where.append("on_pilot_2 = %s")
		vals.append(cint(on_pilot_2))
	if str(on_im) in ("0", "1"):
		where.append("on_im = %s")
		vals.append(cint(on_im))
	if imei:
		where.append("ifnull(imei, '') like %s")
		vals.append(like(imei))
	if vehicle:
		where.append("(ifnull(plate, '') like %s or ifnull(erp_vehicle, '') like %s)")
		vals += [like(vehicle), like(vehicle)]
	if customer:
		where.append("(ifnull(customer_name, '') like %s or ifnull(customer, '') like %s)")
		vals += [like(customer), like(customer)]
	if sim_status:
		where.append("ifnull(sim_status, '') = %s")
		vals.append(sim_status)

	def date_range(col, from_, to_):
		# "to" means the END of that calendar day, not midnight at its start --
		# a person picking "to: 2026-09-05" means "including the 5th".
		if from_:
			where.append("%s >= %%s" % col)
			vals.append(frappe.utils.getdate(from_))
		if to_:
			where.append("%s < %%s" % col)
			vals.append(frappe.utils.getdate(to_) + timedelta(days=1))

	date_range("pilot_last_seen", pilot_last_seen_from, pilot_last_seen_to)
	date_range("im_last_seen", im_last_seen_from, im_last_seen_to)
	date_range("sim_last_seen", sim_last_seen_from, sim_last_seen_to)

	if search:
		where.append(
			"(ifnull(imei,'') like %s or ifnull(plate,'') like %s or ifnull(erp_vehicle,'') like %s"
			" or ifnull(customer_name,'') like %s or ifnull(pilot_vehicle,'') like %s"
			" or ifnull(im_vehicle,'') like %s)"
		)
		vals += [like(search)] * 6

	clause = (" where " + " and ".join(where)) if where else ""

	total = cint(
		frappe.db.sql("select count(*) from `tab%s`%s" % (SNAPSHOT_DT, clause), vals)[0][0]
	)

	fields = ("name, imei, erp_vehicle, plate, customer, customer_name, erp_status, "
	          "server_type, expects_pilot, expects_im, on_pilot, on_pilot_1, on_pilot_2, on_im, "
	          "pilot_checked, im_checked, pilot_active, pilot_last_seen, pilot_msisdn, "
	          "pilot_vehicle, pilot_folder, pilot_account, "
	          "im_vehicle, im_company, im_status, im_last_seen, "
	          "sim_status, sim_last_seen, sim_msisdn, "
	          "verdict, issues")

	rows = frappe.db.sql(
		"select %s from `tab%s`%s order by verdict asc, erp_status asc, imei asc limit %%s offset %%s"
		% (fields, SNAPSHOT_DT, clause),
		vals + [cint(limit) or 100, cint(start)],
		as_dict=True,
	)

	return {"rows": rows, "total": total, "start": cint(start), "limit": cint(limit) or 100}


@frappe.whitelist()
def get_filter_options() -> dict:
	"""The values each column filter can offer, taken from the snapshot itself.

	Read from the data rather than hard-coded: a verdict or an ERP status that
	never occurs should not sit in a dropdown offering zero results.
	"""
	frappe.only_for(READ_ROLES)

	def distinct(col):
		return [
			r[0] for r in frappe.db.sql(
				"select distinct `%s` from `tab%s` where ifnull(`%s`,'') <> '' order by `%s`"
				% (col, SNAPSHOT_DT, col, col)
			)
		]

	return {
		"verdicts": distinct("verdict"),
		"erp_statuses": distinct("erp_status"),
		"sim_statuses": distinct("sim_status"),
	}


@frappe.whitelist()
def get_progress() -> dict:
	"""Whether a run is in flight -- for a page opened mid-run, which gets no
	realtime events it was not around to hear."""
	frappe.only_for(READ_ROLES)

	from frappe.utils.background_jobs import get_jobs

	try:
		queued = get_jobs(site=frappe.local.site, queue="long") or {}
		names = queued.get(frappe.local.site, []) or []
		running = any("fleet_audit" in str(n) for n in names)
	except Exception:
		running = False

	return {"running": running}
